"""
secure_app.py - Google Workspace MCP Server
Uses FastMCP with Streamable HTTP transport + Custom OAuth Proxy

Architecture:
- FastMCP for MCP server with Streamable HTTP transport
- Starlette for OAuth proxy endpoints
- Token validation for @singlefile.io domain restriction
"""

import os
import sys
import logging
import io
import datetime
import time
import secrets
import hashlib
import base64
import httpx
from urllib.parse import urlencode
from typing import Optional, Dict
from contextlib import asynccontextmanager

# FastMCP - the modern way
from mcp.server.fastmcp import FastMCP
from mcp.server.transport_security import TransportSecuritySettings

# Starlette for OAuth routes
from starlette.applications import Starlette
from starlette.routing import Route, Mount
from starlette.requests import Request
from starlette.responses import JSONResponse, RedirectResponse

# Google APIs
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# File handlers
from pypdf import PdfReader
import openpyxl

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("google_mcp")

# Environment variables
KEY_PATH = os.environ.get("SERVICE_ACCOUNT_KEY_PATH", "/app/service-account.json")
ALLOWED_DOMAIN = os.environ.get("ALLOWED_DOMAIN", "singlefile.io")
CLIENT_ID = os.environ.get("CLIENT_ID")
CLIENT_SECRET = os.environ.get("CLIENT_SECRET")
PUBLIC_URL = os.environ.get("PUBLIC_URL", "").rstrip("/")

# Google API Scopes
SCOPES = [
    'https://www.googleapis.com/auth/drive.readonly',
    'https://www.googleapis.com/auth/spreadsheets.readonly',
    'https://www.googleapis.com/auth/documents.readonly',
    'https://www.googleapis.com/auth/presentations.readonly',
    'https://www.googleapis.com/auth/forms.body.readonly',
    'https://www.googleapis.com/auth/calendar.readonly',

    'https://www.googleapis.com/auth/drive.activity.readonly',
]

# =============================================================================
# 2. OAUTH STATE STORAGE
# =============================================================================

auth_states = {}
registered_clients = {}

def cleanup_expired_states():
    now = time.time()
    expired = [k for k, v in auth_states.items() if v.get("expires", 0) < now]
    for k in expired:
        auth_states.pop(k, None)

# =============================================================================
# 3. GOOGLE SERVICE BUILDERS
# =============================================================================

def get_service_account_credentials():
    """Get service account credentials (no impersonation - only accesses shared resources)."""
    creds = service_account.Credentials.from_service_account_file(KEY_PATH, scopes=SCOPES)
    return creds

def build_drive_service():
    return build('drive', 'v3', credentials=get_service_account_credentials())

def build_sheets_service():
    return build('sheets', 'v4', credentials=get_service_account_credentials())

def build_docs_service():
    return build('docs', 'v1', credentials=get_service_account_credentials())

def build_slides_service():
    return build('slides', 'v1', credentials=get_service_account_credentials())

def build_forms_service():
    return build('forms', 'v1', credentials=get_service_account_credentials())

def build_calendar_service():
    return build('calendar', 'v3', credentials=get_service_account_credentials())



def build_activity_service():
    return build('driveactivity', 'v2', credentials=get_service_account_credentials())

# =============================================================================
# 4. FILE CONTENT EXTRACTION
# =============================================================================

def extract_file_content(drive_service, file_id: str, mime_type: str) -> str:
    try:
        if mime_type == 'application/vnd.google-apps.document':
            content = drive_service.files().export(fileId=file_id, mimeType='text/plain').execute()
            return content.decode('utf-8') if isinstance(content, bytes) else content
        elif mime_type == 'application/vnd.google-apps.spreadsheet':
            content = drive_service.files().export(fileId=file_id, mimeType='text/csv').execute()
            return content.decode('utf-8') if isinstance(content, bytes) else content
        elif mime_type == 'application/vnd.google-apps.presentation':
            content = drive_service.files().export(fileId=file_id, mimeType='text/plain').execute()
            return content.decode('utf-8') if isinstance(content, bytes) else content
        else:
            request = drive_service.files().get_media(fileId=file_id)
            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            buffer.seek(0)
            
            if mime_type == 'application/pdf':
                reader = PdfReader(buffer)
                return "\n".join(page.extract_text() or "" for page in reader.pages)
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
                wb = openpyxl.load_workbook(buffer, data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        lines.append(",".join(str(c) if c else "" for c in row))
                return "\n".join(lines)
            elif mime_type.startswith('text/'):
                return buffer.read().decode('utf-8', errors='ignore')
            else:
                return f"[Cannot extract text from {mime_type}]"
    except Exception as e:
        logger.error(f"Content extraction error: {e}")
        return f"[Error extracting content: {e}]"

# =============================================================================
# 5. FASTMCP SERVER WITH TOOLS
# =============================================================================

mcp = FastMCP(
    "Google Workspace MCP",
    stateless_http=False,
    json_response=True,
    transport_security=TransportSecuritySettings(
        enable_dns_rebinding_protection=False,
    ),
)

# -----------------------------------------------------------------------------
# DRIVE TOOLS
# -----------------------------------------------------------------------------

@mcp.tool()
def search_drive(query: str) -> str:
    """Search Google Drive files by content (full-text search)."""
    drive = build_drive_service()
    all_results = []
    
    my_drive_results = drive.files().list(
        q=f"fullText contains '{query}'",
        fields="files(id, name, mimeType, modifiedTime, webViewLink)",
        pageSize=20,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute().get('files', [])
    all_results.extend(my_drive_results)
    
    shared_drives = drive.drives().list(pageSize=50, fields="drives(id, name)").execute().get('drives', [])
    for sd in shared_drives:
        try:
            sd_results = drive.files().list(
                q=f"fullText contains '{query}'",
                driveId=sd['id'],
                corpora="drive",
                fields="files(id, name, mimeType, modifiedTime, webViewLink)",
                pageSize=20,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True
            ).execute().get('files', [])
            for f in sd_results:
                f['_drive_name'] = sd['name']
            all_results.extend(sd_results)
        except Exception as e:
            logger.warning(f"Error searching shared drive {sd['name']}: {e}")
    
    if not all_results:
        return "No files found."
    
    lines = []
    for f in all_results:
        drive_label = f" [Shared Drive: {f['_drive_name']}]" if '_drive_name' in f else ""
        lines.append(f"• {f['name']} ({f['mimeType']}){drive_label}\n  ID: {f['id']}\n  Modified: {f.get('modifiedTime', 'N/A')}\n  Link: {f.get('webViewLink', 'N/A')}")
    return "\n".join(lines)

@mcp.tool()
def search_by_name(name: str) -> str:
    """Search Google Drive files by filename only (faster than full-text search)."""
    drive = build_drive_service()
    all_results = []
    
    my_drive_results = drive.files().list(
        q=f"name contains '{name}'",
        fields="files(id, name, mimeType, modifiedTime, webViewLink)",
        pageSize=20,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute().get('files', [])
    all_results.extend(my_drive_results)
    
    shared_drives = drive.drives().list(pageSize=50, fields="drives(id, name)").execute().get('drives', [])
    for sd in shared_drives:
        try:
            sd_results = drive.files().list(
                q=f"name contains '{name}'",
                driveId=sd['id'],
                corpora="drive",
                fields="files(id, name, mimeType, modifiedTime, webViewLink)",
                pageSize=20,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True
            ).execute().get('files', [])
            for f in sd_results:
                f['_drive_name'] = sd['name']
            all_results.extend(sd_results)
        except Exception as e:
            logger.warning(f"Error searching shared drive {sd['name']}: {e}")
    
    if not all_results:
        return "No files found matching that name."
    
    lines = []
    for f in all_results:
        drive_label = f" [Shared Drive: {f['_drive_name']}]" if '_drive_name' in f else ""
        lines.append(f"• {f['name']} ({f['mimeType']}){drive_label}\n  ID: {f['id']}\n  Modified: {f.get('modifiedTime', 'N/A')}\n  Link: {f.get('webViewLink', 'N/A')}")
    return "\n".join(lines)

@mcp.tool()
def list_recent_files(max_results: int = 20) -> str:
    """List recently modified files across all accessible drives."""
    drive = build_drive_service()
    all_results = []
    
    my_drive_results = drive.files().list(
        q="trashed=false",
        orderBy="modifiedTime desc",
        fields="files(id, name, mimeType, modifiedTime, webViewLink)",
        pageSize=max_results,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute().get('files', [])
    all_results.extend(my_drive_results)
    
    shared_drives = drive.drives().list(pageSize=50, fields="drives(id, name)").execute().get('drives', [])
    for sd in shared_drives:
        try:
            sd_results = drive.files().list(
                q="trashed=false",
                orderBy="modifiedTime desc",
                driveId=sd['id'],
                corpora="drive",
                fields="files(id, name, mimeType, modifiedTime, webViewLink)",
                pageSize=max_results,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True
            ).execute().get('files', [])
            for f in sd_results:
                f['_drive_name'] = sd['name']
            all_results.extend(sd_results)
        except Exception as e:
            logger.warning(f"Error listing shared drive {sd['name']}: {e}")
    
    all_results.sort(key=lambda x: x.get('modifiedTime', ''), reverse=True)
    all_results = all_results[:max_results]
    
    if not all_results:
        return "No recent files found."
    
    lines = ["📅 RECENTLY MODIFIED FILES:\n"]
    for f in all_results:
        drive_label = f" [Shared Drive: {f['_drive_name']}]" if '_drive_name' in f else ""
        icon = "📁" if f['mimeType'] == 'application/vnd.google-apps.folder' else "📄"
        lines.append(f"{icon} {f['name']}{drive_label}\n   Modified: {f.get('modifiedTime', 'N/A')}\n   ID: {f['id']}")
    return "\n".join(lines)

@mcp.tool()
def get_file_info(file_id: str) -> str:
    """Get detailed metadata about a file including size, owner, and sharing info."""
    drive = build_drive_service()
    meta = drive.files().get(
        fileId=file_id,
        fields="id, name, mimeType, size, createdTime, modifiedTime, owners, lastModifyingUser, webViewLink, permissions",
        supportsAllDrives=True
    ).execute()
    
    lines = [f"📄 FILE INFO: {meta.get('name', 'Unknown')}\n"]
    lines.append(f"ID: {meta.get('id')}")
    lines.append(f"Type: {meta.get('mimeType')}")
    if meta.get('size'):
        size_mb = int(meta['size']) / (1024 * 1024)
        lines.append(f"Size: {size_mb:.2f} MB ({meta['size']} bytes)")
    lines.append(f"Created: {meta.get('createdTime', 'N/A')}")
    lines.append(f"Modified: {meta.get('modifiedTime', 'N/A')}")
    
    if meta.get('owners'):
        owners = [o.get('displayName', o.get('emailAddress', 'Unknown')) for o in meta['owners']]
        lines.append(f"Owner(s): {', '.join(owners)}")
    
    if meta.get('lastModifyingUser'):
        lmu = meta['lastModifyingUser']
        lines.append(f"Last modified by: {lmu.get('displayName', lmu.get('emailAddress', 'Unknown'))}")
    
    lines.append(f"Link: {meta.get('webViewLink', 'N/A')}")
    
    if meta.get('permissions'):
        lines.append("\n📤 SHARING:")
        for p in meta['permissions']:
            role = p.get('role', 'unknown')
            ptype = p.get('type', 'unknown')
            if ptype == 'user':
                lines.append(f"  • {p.get('emailAddress', 'Unknown')} ({role})")
            elif ptype == 'group':
                lines.append(f"  • Group: {p.get('emailAddress', 'Unknown')} ({role})")
            elif ptype == 'domain':
                lines.append(f"  • Domain: {p.get('domain', 'Unknown')} ({role})")
            elif ptype == 'anyone':
                lines.append(f"  • Anyone with link ({role})")
    
    return "\n".join(lines)

@mcp.tool()
def get_sharing_permissions(file_id: str) -> str:
    """Get detailed sharing permissions for a file."""
    drive = build_drive_service()
    permissions = drive.permissions().list(
        fileId=file_id,
        fields="permissions(id, type, role, emailAddress, domain, displayName, expirationTime)",
        supportsAllDrives=True
    ).execute().get('permissions', [])
    
    if not permissions:
        return "No sharing permissions found (file may be private)."
    
    lines = ["📤 SHARING PERMISSIONS:\n"]
    for p in permissions:
        role = p.get('role', 'unknown')
        ptype = p.get('type', 'unknown')
        expiry = f" (expires: {p['expirationTime']})" if p.get('expirationTime') else ""
        
        if ptype == 'user':
            name = p.get('displayName', p.get('emailAddress', 'Unknown'))
            email = p.get('emailAddress', '')
            lines.append(f"• User: {name} <{email}> - {role}{expiry}")
        elif ptype == 'group':
            lines.append(f"• Group: {p.get('emailAddress', 'Unknown')} - {role}{expiry}")
        elif ptype == 'domain':
            lines.append(f"• Domain: {p.get('domain', 'Unknown')} - {role}{expiry}")
        elif ptype == 'anyone':
            lines.append(f"• Anyone with link - {role}{expiry}")
    
    return "\n".join(lines)

@mcp.tool()
def get_file_activity(file_id: str, max_results: int = 10) -> str:
    """Get recent activity history for a file (views, edits, comments, shares)."""
    activity = build_activity_service()
    
    try:
        results = activity.activity().query(body={
            "itemName": f"items/{file_id}",
            "pageSize": max_results
        }).execute()
        
        activities = results.get('activities', [])
        if not activities:
            return "No recent activity found for this file."
        
        lines = ["📊 FILE ACTIVITY:\n"]
        for act in activities:
            timestamp = act.get('timestamp', 'Unknown time')
            actors = act.get('actors', [])
            actor_names = []
            for a in actors:
                if 'user' in a:
                    user = a['user'].get('knownUser', {})
                    actor_names.append(user.get('personName', 'Unknown user'))
            
            action = act.get('primaryActionDetail', {})
            action_type = list(action.keys())[0] if action else 'unknown'
            
            action_desc = {
                'create': '📝 Created',
                'edit': '✏️ Edited',
                'move': '📁 Moved',
                'rename': '🏷️ Renamed',
                'delete': '🗑️ Deleted',
                'restore': '♻️ Restored',
                'permissionChange': '🔐 Permissions changed',
                'comment': '💬 Commented',
                'dlpChange': '🔒 DLP change',
                'reference': '🔗 Referenced',
            }.get(action_type, f'📋 {action_type}')
            
            actors_str = ', '.join(actor_names) if actor_names else 'Unknown'
            lines.append(f"• {action_desc} by {actors_str}\n  Time: {timestamp}")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error getting activity: {e}"

@mcp.tool()
def list_file_revisions(file_id: str) -> str:
    """List all revisions (version history) of a file."""
    drive = build_drive_service()
    
    try:
        revisions = drive.revisions().list(
            fileId=file_id,
            fields="revisions(id, modifiedTime, lastModifyingUser, size)",
            pageSize=50
        ).execute().get('revisions', [])
        
        if not revisions:
            return "No revision history available for this file."
        
        lines = [f"📜 VERSION HISTORY ({len(revisions)} revisions):\n"]
        for i, rev in enumerate(revisions, 1):
            user = rev.get('lastModifyingUser', {})
            user_name = user.get('displayName', user.get('emailAddress', 'Unknown'))
            modified = rev.get('modifiedTime', 'Unknown')
            size = rev.get('size', 'N/A')
            if size != 'N/A':
                size = f"{int(size) / 1024:.1f} KB"
            
            lines.append(f"• v{i} - {modified}\n  By: {user_name}\n  Size: {size}\n  Revision ID: {rev['id']}")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error getting revisions: {e}. Note: Google Docs/Sheets/Slides don't expose revision API."

@mcp.tool()
def list_comments(file_id: str) -> str:
    """List all comments on a file."""
    drive = build_drive_service()
    
    try:
        comments = drive.comments().list(
            fileId=file_id,
            fields="comments(id, content, author, createdTime, resolved, replies)",
            pageSize=50
        ).execute().get('comments', [])
        
        if not comments:
            return "No comments found on this file."
        
        lines = [f"💬 COMMENTS ({len(comments)}):\n"]
        for c in comments:
            author = c.get('author', {})
            author_name = author.get('displayName', 'Unknown')
            created = c.get('createdTime', 'Unknown')
            content = c.get('content', '(no content)')
            resolved = "✅ Resolved" if c.get('resolved') else "🔵 Open"
            
            lines.append(f"• {author_name} ({created}) [{resolved}]\n  \"{content}\"")
            
            replies = c.get('replies', [])
            for r in replies:
                reply_author = r.get('author', {}).get('displayName', 'Unknown')
                reply_content = r.get('content', '(no content)')
                lines.append(f"    ↳ {reply_author}: \"{reply_content}\"")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error getting comments: {e}"

@mcp.tool()
def read_file(file_id: str) -> str:
    """Read content from a Google Drive file by its ID (supports shared drives)."""
    drive = build_drive_service()
    meta = drive.files().get(fileId=file_id, fields="name,mimeType", supportsAllDrives=True).execute()
    content = extract_file_content(drive, file_id, meta['mimeType'])
    return f"# {meta['name']}\n\n{content}"

@mcp.tool()
def list_folder(folder_id: str = "root") -> str:
    """List files in a Google Drive folder (supports shared drives). Use 'root' for all accessible files."""
    drive = build_drive_service()
    
    if folder_id == "root":
        all_files = []
        
        my_drive_results = drive.files().list(
            q="'root' in parents and trashed=false",
            fields="files(id, name, mimeType)",
            pageSize=50,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True
        ).execute().get('files', [])
        all_files.extend(my_drive_results)
        
        shared_drives = drive.drives().list(pageSize=50, fields="drives(id, name)").execute().get('drives', [])
        for sd in shared_drives:
            sd_files = drive.files().list(
                q=f"'{sd['id']}' in parents and trashed=false",
                driveId=sd['id'],
                corpora="drive",
                fields="files(id, name, mimeType)",
                pageSize=50,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True
            ).execute().get('files', [])
            for f in sd_files:
                f['_drive_name'] = sd['name']
            all_files.extend(sd_files)
        
        if not all_files:
            return "No files found. Make sure files/drives are shared with the service account."
        
        lines = []
        for f in all_files:
            drive_label = f" [Shared Drive: {f['_drive_name']}]" if '_drive_name' in f else ""
            lines.append(f"• {f['name']} ({f['mimeType']}) - {f['id']}{drive_label}")
        return "\n".join(lines)
    
    results = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id, name, mimeType)",
        pageSize=50,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute().get('files', [])
    
    if not results:
        return "Folder is empty."
    
    return "\n".join([f"• {f['name']} ({f['mimeType']}) - {f['id']}" for f in results])

@mcp.tool()
def list_shared_drives() -> str:
    """List all shared drives accessible to the service account."""
    drive = build_drive_service()
    results = drive.drives().list(pageSize=50, fields="drives(id, name)").execute().get('drives', [])
    
    if not results:
        return "No shared drives found. Make sure the service account is a member of the shared drive."
    
    return "\n".join([f"• {d['name']} (ID: {d['id']})" for d in results])

@mcp.tool()
def list_shared_drive_members(drive_id: str) -> str:
    """List all members/permissions of a shared drive."""
    drive = build_drive_service()
    
    try:
        permissions = drive.permissions().list(
            fileId=drive_id,
            supportsAllDrives=True,
            fields="permissions(id, type, role, emailAddress, displayName)"
        ).execute().get('permissions', [])
        
        if not permissions:
            return "No members found or unable to access shared drive permissions."
        
        lines = ["👥 SHARED DRIVE MEMBERS:\n"]
        for p in permissions:
            role = p.get('role', 'unknown')
            ptype = p.get('type', 'unknown')
            name = p.get('displayName', p.get('emailAddress', 'Unknown'))
            email = p.get('emailAddress', '')
            
            role_emoji = {'organizer': '👑', 'fileOrganizer': '📁', 'writer': '✏️', 'commenter': '💬', 'reader': '👁️'}.get(role, '👤')
            
            if ptype == 'user':
                lines.append(f"{role_emoji} {name} <{email}> - {role}")
            elif ptype == 'group':
                lines.append(f"{role_emoji} Group: {email} - {role}")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error listing shared drive members: {e}"

# -----------------------------------------------------------------------------
# DOCUMENT TOOLS
# -----------------------------------------------------------------------------

@mcp.tool()
def read_spreadsheet(spreadsheet_id: str, range: str = "A1:Z1000") -> str:
    """Read data from a Google Spreadsheet."""
    sheets = build_sheets_service()
    result = sheets.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=range).execute()
    values = result.get('values', [])
    if not values:
        return "Spreadsheet is empty."
    return "\n".join([",".join(str(c) for c in row) for row in values])

@mcp.tool()
def read_document(document_id: str) -> str:
    """Read content from a Google Doc."""
    docs = build_docs_service()
    doc = docs.documents().get(documentId=document_id).execute()
    content = []
    for elem in doc.get('body', {}).get('content', []):
        if 'paragraph' in elem:
            for el in elem['paragraph'].get('elements', []):
                if 'textRun' in el:
                    content.append(el['textRun'].get('content', ''))
    return f"# {doc.get('title', 'Untitled')}\n\n{''.join(content)}"

@mcp.tool()
def read_presentation(presentation_id: str) -> str:
    """Read content from a Google Slides presentation."""
    slides = build_slides_service()
    pres = slides.presentations().get(presentationId=presentation_id).execute()
    content = [f"# {pres.get('title', 'Untitled Presentation')}"]
    for i, slide in enumerate(pres.get('slides', []), 1):
        content.append(f"\n## Slide {i}")
        for elem in slide.get('pageElements', []):
            if 'shape' in elem and 'text' in elem['shape']:
                for te in elem['shape']['text'].get('textElements', []):
                    if 'textRun' in te:
                        content.append(te['textRun'].get('content', '').strip())
    return "\n".join(content)

@mcp.tool()
def read_form(form_id: str) -> str:
    """Read questions and structure from a Google Form."""
    forms = build_forms_service()
    
    try:
        form = forms.forms().get(formId=form_id).execute()
        
        lines = [f"📋 FORM: {form.get('info', {}).get('title', 'Untitled Form')}\n"]
        
        if form.get('info', {}).get('description'):
            lines.append(f"Description: {form['info']['description']}\n")
        
        items = form.get('items', [])
        if not items:
            lines.append("(No questions in this form)")
        
        for i, item in enumerate(items, 1):
            title = item.get('title', '(Untitled question)')
            
            if 'questionItem' in item:
                q = item['questionItem']['question']
                required = "* " if q.get('required') else ""
                
                if 'choiceQuestion' in q:
                    cq = q['choiceQuestion']
                    q_type = cq.get('type', 'CHOICE')
                    options = [opt.get('value', '') for opt in cq.get('options', [])]
                    lines.append(f"{i}. {required}{title} [{q_type}]")
                    for opt in options:
                        lines.append(f"   ○ {opt}")
                elif 'textQuestion' in q:
                    paragraph = "paragraph" if q['textQuestion'].get('paragraph') else "short"
                    lines.append(f"{i}. {required}{title} [TEXT - {paragraph}]")
                elif 'scaleQuestion' in q:
                    sq = q['scaleQuestion']
                    lines.append(f"{i}. {required}{title} [SCALE: {sq.get('low', 1)}-{sq.get('high', 5)}]")
                elif 'dateQuestion' in q:
                    lines.append(f"{i}. {required}{title} [DATE]")
                elif 'timeQuestion' in q:
                    lines.append(f"{i}. {required}{title} [TIME]")
                elif 'fileUploadQuestion' in q:
                    lines.append(f"{i}. {required}{title} [FILE UPLOAD]")
                else:
                    lines.append(f"{i}. {required}{title}")
            
            elif 'pageBreakItem' in item:
                lines.append(f"\n--- Page Break: {title} ---\n")
            elif 'textItem' in item:
                lines.append(f"\n📝 {title}")
            elif 'imageItem' in item:
                lines.append(f"🖼️ [Image: {title}]")
            elif 'videoItem' in item:
                lines.append(f"🎬 [Video: {title}]")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error reading form: {e}"

@mcp.tool()
def read_form_responses(form_id: str) -> str:
    """Read all responses submitted to a Google Form."""
    forms = build_forms_service()
    
    try:
        form = forms.forms().get(formId=form_id).execute()
        form_title = form.get('info', {}).get('title', 'Untitled Form')
        
        items = form.get('items', [])
        question_map = {}
        for item in items:
            if 'questionItem' in item:
                qid = item['questionItem']['question'].get('questionId')
                if qid:
                    question_map[qid] = item.get('title', 'Unknown question')
        
        responses = forms.forms().responses().list(formId=form_id).execute()
        response_list = responses.get('responses', [])
        
        if not response_list:
            return f"No responses submitted to form: {form_title}"
        
        lines = [f"📊 RESPONSES FOR: {form_title}\nTotal responses: {len(response_list)}\n"]
        
        for i, resp in enumerate(response_list, 1):
            lines.append(f"\n--- Response {i} ---")
            lines.append(f"Submitted: {resp.get('lastSubmittedTime', 'Unknown')}")
            
            answers = resp.get('answers', {})
            for qid, answer_data in answers.items():
                question = question_map.get(qid, f'Question {qid}')
                answer = answer_data.get('textAnswers', {})
                answer_values = [a.get('value', '') for a in answer.get('answers', [])]
                lines.append(f"Q: {question}")
                lines.append(f"A: {', '.join(answer_values)}")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error reading form responses: {e}"

# -----------------------------------------------------------------------------
# CALENDAR TOOLS
# -----------------------------------------------------------------------------

@mcp.tool()
def list_calendars() -> str:
    """List all calendars accessible to the service account."""
    cal = build_calendar_service()
    calendars = cal.calendarList().list().execute().get('items', [])
    if not calendars:
        return "No calendars found. Make sure calendars are shared with the service account."
    return "\n".join([f"• {c['summary']} (ID: {c['id']})" for c in calendars])

@mcp.tool()
def list_events(calendar_id: str = "primary", max_results: int = 10) -> str:
    """List upcoming events from a calendar."""
    cal = build_calendar_service()
    now = datetime.datetime.utcnow().isoformat() + 'Z'
    try:
        events = cal.events().list(
            calendarId=calendar_id,
            timeMin=now,
            maxResults=max_results,
            singleEvents=True,
            orderBy='startTime'
        ).execute().get('items', [])
    except Exception as e:
        return f"Error accessing calendar: {e}. Make sure the calendar is shared with the service account."
    
    if not events:
        return "No upcoming events."
    
    lines = []
    for e in events:
        start = e['start'].get('dateTime', e['start'].get('date'))
        lines.append(f"• {e['summary']} - {start}")
    
    return "\n".join(lines)

@mcp.tool()
def search_calendar_events(calendar_id: str = "primary", query: str = "", time_min: str = "", time_max: str = "", max_results: int = 20) -> str:
    """Search calendar events by keyword and/or date range.
    
    Args:
        calendar_id: Calendar ID (use 'primary' for main calendar)
        query: Search term to find in event titles/descriptions
        time_min: Start of date range (ISO format, e.g., '2024-01-01T00:00:00Z')
        time_max: End of date range (ISO format, e.g., '2024-12-31T23:59:59Z')
        max_results: Maximum number of results to return
    """
    cal = build_calendar_service()
    
    try:
        params = {
            'calendarId': calendar_id,
            'maxResults': max_results,
            'singleEvents': True,
            'orderBy': 'startTime'
        }
        
        if query:
            params['q'] = query
        if time_min:
            params['timeMin'] = time_min
        else:
            params['timeMin'] = datetime.datetime.utcnow().isoformat() + 'Z'
        if time_max:
            params['timeMax'] = time_max
        
        events = cal.events().list(**params).execute().get('items', [])
        
        if not events:
            return "No events found matching your criteria."
        
        lines = [f"📅 CALENDAR EVENTS ({len(events)} found):\n"]
        for e in events:
            start = e['start'].get('dateTime', e['start'].get('date'))
            end = e['end'].get('dateTime', e['end'].get('date'))
            summary = e.get('summary', '(No title)')
            location = e.get('location', '')
            
            lines.append(f"• {summary}")
            lines.append(f"  Start: {start}")
            lines.append(f"  End: {end}")
            if location:
                lines.append(f"  Location: {location}")
            if e.get('description'):
                desc = e['description'][:100] + '...' if len(e.get('description', '')) > 100 else e.get('description', '')
                lines.append(f"  Description: {desc}")
            lines.append("")
        
        return "\n".join(lines)
    except Exception as e:
        return f"Error searching calendar: {e}"



# =============================================================================
# 6. OAUTH PROXY ENDPOINTS
# =============================================================================

async def oauth_well_known(request: Request):
    return JSONResponse({
        "issuer": PUBLIC_URL,
        "authorization_endpoint": f"{PUBLIC_URL}/authorize",
        "token_endpoint": f"{PUBLIC_URL}/token",
        "registration_endpoint": f"{PUBLIC_URL}/register",
        "userinfo_endpoint": f"{PUBLIC_URL}/userinfo",
        "response_types_supported": ["code"],
        "grant_types_supported": ["authorization_code", "refresh_token"],
        "subject_types_supported": ["public"],
        "id_token_signing_alg_values_supported": ["RS256"],
        "scopes_supported": ["openid", "email", "profile"],
        "token_endpoint_auth_methods_supported": ["none", "client_secret_post"],
        "code_challenge_methods_supported": ["S256"],
    })

async def oauth_protected_resource(request: Request):
    return JSONResponse({
        "resource": PUBLIC_URL,
        "authorization_servers": [PUBLIC_URL],
        "scopes_supported": ["openid", "email", "profile"],
        "bearer_methods_supported": ["header"],
    })

async def oauth_authorize(request: Request):
    cleanup_expired_states()
    params = request.query_params
    
    internal_state = secrets.token_urlsafe(32)
    auth_states[internal_state] = {
        "client_id": params.get("client_id", ""),
        "client_state": params.get("state", ""),
        "client_redirect_uri": params.get("redirect_uri", ""),
        "code_challenge": params.get("code_challenge", ""),
        "code_challenge_method": params.get("code_challenge_method", "S256"),
        "scope": params.get("scope", "openid email profile"),
        "expires": time.time() + 600
    }
    
    google_params = {
        "client_id": CLIENT_ID,
        "redirect_uri": f"{PUBLIC_URL}/oauth2callback",
        "response_type": "code",
        "scope": "openid email profile " + " ".join(SCOPES),
        "state": internal_state,
        "access_type": "offline",
        "prompt": "consent",
        "hd": ALLOWED_DOMAIN,
    }
    
    return RedirectResponse(f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(google_params)}")

async def oauth_callback(request: Request):
    params = request.query_params
    code = params.get("code")
    internal_state = params.get("state")
    error = params.get("error")
    
    if error:
        return JSONResponse({"error": error}, status_code=400)
    
    stored = auth_states.pop(internal_state, None)
    if not stored:
        return JSONResponse({"error": "Invalid or expired state"}, status_code=400)
    
    logger.info("📥 Google callback received, exchanging code...")
    
    async with httpx.AsyncClient() as client:
        resp = await client.post("https://oauth2.googleapis.com/token", data={
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": f"{PUBLIC_URL}/oauth2callback"
        })
        
        if resp.status_code != 200:
            return JSONResponse({"error": "Token exchange failed"}, status_code=400)
        
        token_data = resp.json()
        logger.info(f"✅ Got tokens from Google")
    
    proxy_code = secrets.token_urlsafe(32)
    auth_states[proxy_code] = {
        "tokens": token_data,
        "code_challenge": stored.get("code_challenge"),
        "code_challenge_method": stored.get("code_challenge_method"),
        "expires": time.time() + 300
    }
    
    redirect_url = f"{stored['client_redirect_uri']}?code={proxy_code}&state={stored['client_state']}"
    return RedirectResponse(redirect_url)

async def oauth_token(request: Request):
    form = await request.form()
    grant_type = form.get("grant_type")
    code_verifier = form.get("code_verifier")
    
    logger.info(f"📥 Token request - grant_type: {grant_type}")
    
    if grant_type == "authorization_code":
        code = form.get("code")
        stored = auth_states.pop(code, None)
        
        if not stored or stored.get("expires", 0) < time.time():
            return JSONResponse({"error": "invalid_grant"}, status_code=400)
        
        if stored.get("code_challenge") and code_verifier:
            verifier_hash = hashlib.sha256(code_verifier.encode()).digest()
            computed_challenge = base64.urlsafe_b64encode(verifier_hash).rstrip(b'=').decode()
            if computed_challenge != stored["code_challenge"]:
                return JSONResponse({"error": "invalid_grant"}, status_code=400)
            logger.info("✅ PKCE verification passed")
        
        token_data = stored["tokens"]
        return JSONResponse({
            "access_token": token_data["access_token"],
            "token_type": "bearer",
            "expires_in": token_data.get("expires_in", 3600),
            "refresh_token": token_data.get("refresh_token", ""),
            "scope": token_data.get("scope", ""),
        })
    
    elif grant_type == "refresh_token":
        refresh_token = form.get("refresh_token")
        
        async with httpx.AsyncClient() as client:
            resp = await client.post("https://oauth2.googleapis.com/token", data={
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token"
            })
            
            if resp.status_code != 200:
                return JSONResponse({"error": "invalid_grant"}, status_code=400)
            
            token_data = resp.json()
            return JSONResponse({
                "access_token": token_data["access_token"],
                "token_type": "bearer",
                "expires_in": token_data.get("expires_in", 3600),
                "refresh_token": token_data.get("refresh_token", refresh_token),
                "scope": token_data.get("scope", ""),
            })
    
    return JSONResponse({"error": "unsupported_grant_type"}, status_code=400)

async def oauth_register(request: Request):
    try:
        body = await request.json()
    except:
        return JSONResponse({"error": "invalid_request"}, status_code=400)
    
    client_id = secrets.token_urlsafe(24)
    client_secret = secrets.token_urlsafe(32)
    
    registered_clients[client_id] = {
        "client_secret": client_secret,
        "redirect_uris": body.get("redirect_uris", []),
        "client_name": body.get("client_name", "Unknown"),
    }
    
    logger.info(f"✅ Registered client: {body.get('client_name', 'Unknown')}")
    
    return JSONResponse({
        "client_id": client_id,
        "client_secret": client_secret,
        "client_id_issued_at": int(time.time()),
        "client_secret_expires_at": 0,
        "redirect_uris": body.get("redirect_uris", []),
        "client_name": body.get("client_name", "Unknown"),
        "token_endpoint_auth_method": "none",
        "grant_types": ["authorization_code", "refresh_token"],
        "response_types": ["code"],
    }, status_code=201)

async def health_check(request: Request):
    return JSONResponse({"status": "healthy", "timestamp": datetime.datetime.utcnow().isoformat()})

# =============================================================================
# 7. COMBINED STARLETTE APP
# =============================================================================

@asynccontextmanager
async def lifespan(app: Starlette):
    async with mcp.session_manager.run():
        logger.info("🚀 MCP session manager started")
        yield
    logger.info("👋 MCP session manager stopped")

mcp_app = mcp.streamable_http_app()

app = Starlette(
    routes=[
        Route("/.well-known/oauth-authorization-server", endpoint=oauth_well_known),
        Route("/.well-known/oauth-protected-resource", endpoint=oauth_protected_resource),
        Route("/.well-known/oauth-protected-resource/mcp", endpoint=oauth_protected_resource),
        Route("/.well-known/openid-configuration", endpoint=oauth_well_known),
        Route("/authorize", endpoint=oauth_authorize, methods=["GET"]),
        Route("/oauth2callback", endpoint=oauth_callback, methods=["GET"]),
        Route("/token", endpoint=oauth_token, methods=["POST"]),
        Route("/register", endpoint=oauth_register, methods=["POST"]),
        Route("/health", endpoint=health_check),
        Mount("/", app=mcp_app),
    ],
    lifespan=lifespan,
)

# =============================================================================
# 8. ENTRYPOINT
# =============================================================================

if __name__ == "__main__":
    import uvicorn
    
    if not PUBLIC_URL:
        logger.error("❌ PUBLIC_URL required!")
        sys.exit(1)
    if not CLIENT_ID or not CLIENT_SECRET:
        logger.error("❌ CLIENT_ID and CLIENT_SECRET required!")
        sys.exit(1)
    
    logger.info(f"🚀 Starting Google Workspace MCP Server")
    logger.info(f"   PUBLIC_URL: {PUBLIC_URL}")
    logger.info(f"   MCP endpoint: {PUBLIC_URL}/mcp")
    
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))